import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def exporter_excel(colonnes, donnees, nom_fichier):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VentePro Export"

    # Style entête
    entete_font = Font(bold=True, color="FFFFFF")
    entete_fill = PatternFill("solid", fgColor="E65100")
    entete_align = Alignment(horizontal="center")

    for col_idx, col_nom in enumerate(colonnes, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_nom)
        cell.font = entete_font
        cell.fill = entete_fill
        cell.alignment = entete_align

    # Données
    for row_idx, row in enumerate(donnees, 2):
        for col_idx, valeur in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=valeur)

    # Largeur colonnes auto
    for col in ws.columns:
        max_width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_width + 4

    wb.save(nom_fichier)
    return nom_fichier